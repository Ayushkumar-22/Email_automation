#!/usr/bin/env python
# coding: utf-8

# In[1]:


import win32com.client as wc
import pandas as pd


# In[2]:


import os


# In[43]:


#downloading mail from outlook n .msg
import win32com.client

def download_emails_by_subject(subject_keyword, num_emails_to_download=5):
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)  # 6 corresponds to the Inbox folder

    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)  # Sort by received time, descending

    downloaded_count = 0

    for message in messages:
        if subject_keyword.lower() in message.Subject.lower():
            # Save the email as .msg file
            msg_file_path = f"C:\\Users\\yogen\\OneDrive\\Desktop\\New folder\\mail.msg"
            message.SaveAs(msg_file_path, 3)  # 3 corresponds to olMsg (olMsg is the default value)
            
            print(f"Saved: {msg_file_path}")

            downloaded_count += 1

            if downloaded_count == num_emails_to_download:
                break

    print(f"Downloaded {downloaded_count} emails with subject containing '{subject_keyword}'.")

# Replace 'YourSubjectKeyword' with the actual subject you are looking for
download_emails_by_subject('e-Invoice errors on date', num_emails_to_download=3)


# In[3]:


#converting mail in xlsx.

msg_location = r'C:\Users\yogen\OneDrive\Desktop\mail automation\input'
output_location =r'C:\Users\yogen\OneDrive\Desktop\mail automation\output'
files=os.listdir(msg_location)
for file in files:
    if file.endswith(".msg"):
        outlook=wc.Dispatch('outlook.Application').GetNamespace('MAPI')
        msg=outlook.OpenSharedItem(msg_location+'/'+file)
        html_str=msg.HTMLBody
        try:
            pd.read_html(html_str)[0].to_csv(output_location+'\\'+file[:-4]+'.csv',index=False)
        except ValueError:
            continue


# In[7]:


#vlookup with master data

# import pandas 
import pandas as pd 
from pathlib import Path
   
# read csv data 
df1 = pd.read_csv(r"C:\Users\yogen\OneDrive\Desktop\mail automation\output\mail.csv",header=0,skiprows=1) 
df2 = pd.read_excel(r"C:\Users\yogen\OneDrive\Desktop\mail automation\master.xlsx",header=0,skiprows=1) 
df3 = df2[["TRADE NAME / LEGAL NAME", "email id"]]

   
inner_join = pd.merge(df1,
                      df3,  
                      on ='TRADE NAME / LEGAL NAME', 
#                       right_on ='email id',
                      how ='inner') 
inner_join_1=inner_join.drop_duplicates()

inner_join_1


# In[ ]:





# In[4]:


# for create new excel file and add all the data into a new excel file
# Import the necessary libraries
import openpyxl
import pandas as pd
import os

os.chdir(r'C:\Users\yogen\OneDrive\Desktop\mail automation\output')

# Create a new Excel workbook
workbook = openpyxl.Workbook()
# Select the default sheet (usually named 'Sheet')
sheet = workbook.active
# Add data to the Excel sheet
data = inner_join_1
for row in data:
       # sheet.append(row)
    # Save the workbook to a file
    workbook.save("my_excel_file.xlsx")
    # Print a success message
#     print("Excel file created successfully!")
inner_join_1.to_excel(r"C:\Users\yogen\OneDrive\Desktop\mail automation\output\my_excel_file.xlsx",index=False)


# In[ ]:





# In[5]:


# filtering data of 1 mail id
import win32com.client as wc
import csv
import pandas as pd
import uuid
# import jpype
# import asposecells

data = pd.read_excel(r"C:\Users\yogen\OneDrive\Desktop\mail automation\output\my_excel_file.xlsx")

# Get unique email addresses
unique_emails = data['email id'].unique()

# Iterate over unique email addresses
for email in unique_emails:
    # Filter data for the current email address
    filtered_data = data[data['email id'] == email]
    unique_id=uuid.uuid4()
    # Define the output Excel file path for the current email address
#     output_excel_path = fr"C:\Users\yogen\OneDrive\Desktop\mail automation\final data\final_data.xlsx"
    output_excel_path=r"C:\Users\yogen\OneDrive\Desktop\mail automation\final data\final_data_"+str(unique_id)+".xlsx"
    # Save the filtered data to an Excel file
    filtered_data.to_excel(output_excel_path, index=False)
    
    # Print information about the saved file
    print(filtered_data)
    import os
    import win32com.client as wc
    from PIL import ImageGrab

    # Define the path to the Excel workbook
    workbook_path =output_excel_path
    workbook_path
    # Create an instance of Excel
    excel = wc.Dispatch('excel.application')
    wb = excel.Workbooks.Open(workbook_path)
    sheet = wb.Sheets.Item(1)
    sheet.Columns.AutoFit()
    
    # end_row=sheet.max_row
    # end_column=sheet.max_column
    # Copy the range from Excel to clipboard
    copyrange = sheet.Range("A1:G3")
    copyrange.Select()
    copyrange.CopyPicture(Appearance=1, Format=2)
    excel.Selection.Copy()
 
    
    # Save the clipboard content as an image file
    image_path = r"C:\Users\yogen\OneDrive\Desktop\mail automation\final data\final_data_"+str(unique_id)+".png"
    print(image_path)
    ImageGrab.grabclipboard().save(image_path)

    # Define the HTML body with the correct image source
    html_body = f"""
        <div>
            Please review the following report and respond with your feedback.
        </div>
        <div>
            <img src="cid:myimage">
        </div>
    """

    # Create an instance of Outlook
    outlook = wc.Dispatch('Outlook.Application')

    # Create a new email message
    message = outlook.CreateItem(0)

    # Set email properties
    message.To = email
    message.Subject = 'Please review!'
    message.HTMLBody = html_body

    # Attach the image as an embedded attachment
    image_attachment = message.Attachments.Add(Source=image_path, Type=1, Position=0, DisplayName="myimage")

    # Set the content ID (CID) for the embedded image
    image_attachment.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001E", "myimage")

    # Display the email message
    message.Display()
    import psutil

  



# In[ ]:




